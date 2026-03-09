import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_DIR = Path("outputs/csv")
EXCEL_DIR = Path("outputs/excel")
SHEET_SOURCES = {
    "KRW시나리오": "kwrw_stablecoin_scenario_*.csv",
    "스크리닝스코어카드": "screening_scorecard_*.csv",
    "거시지표스냅샷": "macro_snapshot_*.csv",
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1565C0")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SCORE_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
SCORE_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
SCORE_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
INTEGER_RE = re.compile(r"^-?\d+$")
FLOAT_RE = re.compile(r"^-?\d+\.\d+$")


def _latest_csv(pattern: str) -> Path:
    files = sorted(CSV_DIR.glob(pattern), reverse=True)
    if not files:
        raise FileNotFoundError(f"CSV 파일을 찾을 수 없습니다: {CSV_DIR / pattern}")
    return files[0]


def _parse_value(raw: str):
    value = raw.strip()
    if value == "":
        return ""
    if INTEGER_RE.match(value):
        return int(value)
    if FLOAT_RE.match(value):
        return float(value)
    return value


def _apply_header_style(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _apply_number_format(worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.00"


def _apply_score_format(worksheet) -> None:
    for index, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").lower()
        if "score" not in header:
            continue
        column_letter = get_column_letter(index)
        score_range = f"{column_letter}2:{column_letter}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            score_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=SCORE_GREEN),
        )
        worksheet.conditional_formatting.add(
            score_range,
            CellIsRule(operator="between", formula=["50", "69"], fill=SCORE_YELLOW),
        )
        worksheet.conditional_formatting.add(
            score_range,
            CellIsRule(operator="lessThanOrEqual", formula=["49"], fill=SCORE_RED),
        )


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        if not lengths:
            continue
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(lengths) + 2, 40)


def _write_sheet(workbook: Workbook, title: str, csv_path: Path) -> None:
    worksheet = workbook.create_sheet(title=title)
    with csv_path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        for row in reader:
            worksheet.append([_parse_value(value) for value in row])
    _apply_header_style(worksheet)
    _apply_number_format(worksheet)
    _apply_score_format(worksheet)
    _autosize_columns(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def generate_excel() -> str:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, pattern in SHEET_SOURCES.items():
        csv_path = _latest_csv(pattern)
        _write_sheet(workbook, sheet_name, csv_path)

    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.today().strftime("%Y%m%d")
    output_path = EXCEL_DIR / f"danal_research_{date_str}.xlsx"
    workbook.save(output_path)
    return str(output_path)
